# -*- coding: utf-8 -*-
"""
어학연수 지도 — 데이터 반영 스크립트
마스터시트(xlsx)의 내용을 지도 HTML에 심어 넣습니다.

[사용법]
1. 이 파일, 마스터시트, 지도 HTML(map.html)을 같은 폴더에 두기
2. 아래 MASTER / MAP 파일명 확인
3. 실행:  python build_map.py
   → map.html의 데이터가 교체되고, 이전 버전은 map_backup.html로 보관됩니다.

[규칙 — 지도가 자동으로 처리하는 것]
- status가 '운영중'이 아닌 센터는 표시하지 않음 (도시 카운트에서도 제외)
- lat/lng가 빈 센터·기숙사는 목록엔 나오되 지도 점은 생기지 않음
- 같은 좌표가 겹치면 자동으로 살짝 벌려서 표시

[주의]
- map.html 안의 `const D = ` 와 `const CONSULT` 앵커 텍스트를 지우면 안 됩니다.
- 시트 컬럼명(헤더)을 바꾸면 안 됩니다. 컬럼 추가는 괜찮습니다.
"""
import json
import shutil
from openpyxl import load_workbook

# ===== 파일명 확인 =====
MASTER = "어학연수_지도사이트_마스터시트_v3.10.xlsx"
MAP = "index.html"
# ======================

def rows(wb, sheet):
    ws = wb[sheet]
    hdr = [c.value for c in ws[1]]
    out = []
    for r in range(2, ws.max_row + 1):
        d = dict(zip(hdr, [ws.cell(row=r, column=j).value for j in range(1, ws.max_column + 1)]))
        if any(v not in (None, "") for v in d.values()):
            out.append({k: (v if v is not None else "") for k, v in d.items()})
    return out

def main():
    wb = load_workbook(MASTER)
    D = {
        "cities": rows(wb, "cities"),
        "schools": [{k: s.get(k, "") for k in ("school_id", "name_en", "name_ko", "intro", "website")}
                    for s in rows(wb, "schools")],
        "centers": rows(wb, "centers"),
        "residences": rows(wb, "residences"),
    }

    s = open(MAP, encoding="utf-8").read()
    i = s.find("const D = ")
    j = s.find(";\nconst CONSULT")
    if i < 0 or j < 0:
        print("[중단] map.html에서 데이터 앵커(const D / const CONSULT)를 찾지 못했습니다.")
        return

    shutil.copy(MAP, "map_backup.html")
    s = s[:i] + "const D = " + json.dumps(D, ensure_ascii=False) + s[j:]
    open(MAP, "w", encoding="utf-8").write(s)

    no_coord = [x.get("center_id") or x.get("residence_id")
                for x in D["centers"] + D["residences"] if not x.get("lat")]
    print("반영 완료:", {k: len(v) for k, v in D.items()})
    print("지도 미표시(좌표 없음):", no_coord if no_coord else "없음")
    print("이전 버전 백업 → map_backup.html")

if __name__ == "__main__":
    main()
