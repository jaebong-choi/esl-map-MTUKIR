# -*- coding: utf-8 -*-
"""
어학연수 지도 사이트 - 좌표 변환 v3 (영국 + 아일랜드 + 몰타 지원)

[사용법]
1. 이 파일과 '어학연수_지도사이트_마스터시트_v3.7.xlsx'를 같은 폴더에 두기
2. 아래 API_KEY 에 본인 키 붙여넣기 (기존 geocode_v2.py에서 복사하면 됩니다)
3. 그 폴더에서 cmd 열고:  python geocode_v2.py

[결과]
- 마스터시트_v3.5_좌표완료.xlsx : 좌표가 채워진 파일
- 좌표검수리포트_v2.txt        : 실패·의심 건 목록

[참고]
- 이미 좌표가 있는 행은 건너뜁니다 → 이번 변환 대상은 21곳 (더블린·몰타·신규 브랜드)
  — 주소 미확인 2곳(St Giles iQ, EP 더블린 하우스쉐어)은 실패로 나오는 게 정상
- St Giles 브라이튼 iQ는 주소가 '미확인'이라 실패로 나오는 게 정상입니다
"""

import time
import requests
from openpyxl import load_workbook

# ===== 여기만 수정 =====
API_KEY = "여기에 본인 구글 지오코딩 API 키 붙여넣기"  # 공개 리포에 실제 키를 커밋하지 말 것
INPUT_FILE = "어학연수_지도사이트_마스터시트_v3.7.xlsx"
OUTPUT_FILE = "마스터시트_v3.8_좌표완료.xlsx"
# =====================

REPORT_FILE = "좌표검수리포트_v3.txt"
URL = "https://maps.googleapis.com/maps/api/geocode/json"

# 도시별 중심 좌표 (의심 건 판별용)
CITY_COORDS = {
    "uk-london": (51.5074, -0.1278), "uk-cambridge": (52.2053, 0.1218),
    "uk-oxford": (51.7520, -1.2577), "uk-manchester": (53.4808, -2.2426),
    "uk-liverpool": (53.4084, -2.9916), "uk-brighton": (50.8225, -0.1372),
    "uk-bristol": (51.4545, -2.5879), "uk-bath": (51.3811, -2.3590),
    "uk-bournemouth": (50.7192, -1.8808), "uk-edinburgh": (55.9533, -3.1883),
    "uk-torquay": (50.4619, -3.5253), "uk-leeds": (53.8008, -1.5491),
    "uk-eastbourne": (50.7687, 0.2845), "ie-dublin": (53.3498, -6.2603), "mt-stjulians": (35.9186, 14.4892),
    "uk-birmingham": (52.4862, -1.8904),
}
MAX_KM = 25


def distance_km(a, b):
    from math import radians, sin, cos, asin, sqrt
    lat1, lon1, lat2, lon2 = map(radians, [a[0], a[1], b[0], b[1]])
    h = sin((lat2 - lat1) / 2) ** 2 + cos(lat1) * cos(lat2) * sin((lon2 - lon1) / 2) ** 2
    return 6371 * 2 * asin(sqrt(h))


def geocode(address, region):
    r = requests.get(URL, params={"address": address, "key": API_KEY,
                                  "region": region, "language": "en"}, timeout=15)
    d = r.json()
    if d.get("status") != "OK" or not d.get("results"):
        return None, d.get("status", "ERROR"), d.get("error_message", "")
    top = d["results"][0]
    loc = top["geometry"]["location"]
    return (loc["lat"], loc["lng"]), top["geometry"].get("location_type", ""), top.get("formatted_address", "")


def main():
    if API_KEY.startswith("여기에"):
        print("[중단] API_KEY를 먼저 입력하세요.")
        return

    wb = load_workbook(INPUT_FILE)
    report, done, skipped, failed, suspect = [], 0, 0, 0, 0

    for sheet in ("centers", "residences"):
        ws = wb[sheet]
        hdr = {c.value: c.column for c in ws[1]}
        id_col = hdr.get("center_id") or hdr.get("residence_id")

        for row in range(2, ws.max_row + 1):
            rid = ws.cell(row=row, column=id_col).value
            if not rid:
                continue
            if ws.cell(row=row, column=hdr["lat"]).value:
                skipped += 1
                continue

            addr = ws.cell(row=row, column=hdr["address"]).value
            city = str(ws.cell(row=row, column=hdr["city_id"]).value or "")
            if not addr or str(addr).strip() == "미확인":
                report.append(f"[실패] {rid} : 주소 없음(미확인)")
                failed += 1
                continue

            # 국가 인식: ie-* 도시는 아일랜드, 그 외 영국
            if city.startswith("ie-"):   suffix, region = ", Ireland", "ie"
            elif city.startswith("mt-"): suffix, region = ", Malta", "mt"
            else:                        suffix, region = ", UK", "uk"
            up = str(addr).upper()
            query = str(addr) if ("IRELAND" in up or "MALTA" in up or "UK" in up
                                  or "UNITED KINGDOM" in up) else str(addr) + suffix

            try:
                coords, ltype, formatted = geocode(query, region)
            except Exception as e:
                report.append(f"[실패] {rid} : 요청 오류 {e}")
                failed += 1
                continue

            if not coords:
                report.append(f"[실패] {rid} : {ltype} / 주소={addr}")
                failed += 1
                time.sleep(0.1)
                continue

            ws.cell(row=row, column=hdr["lat"], value=round(coords[0], 6))
            ws.cell(row=row, column=hdr["lng"], value=round(coords[1], 6))
            done += 1

            warn = []
            if city in CITY_COORDS:
                dist = distance_km(coords, CITY_COORDS[city])
                if dist > MAX_KM:
                    warn.append(f"도시중심에서 {dist:.1f}km")
            if ltype in ("APPROXIMATE", "GEOMETRIC_CENTER"):
                warn.append(f"정확도 낮음({ltype})")
            if warn:
                report.append(f"[의심] {rid} : {', '.join(warn)}\n"
                              f"        입력={addr}\n        구글인식={formatted}")
                suspect += 1

            print(f"  {rid} → {coords[0]:.5f}, {coords[1]:.5f}")
            time.sleep(0.05)

    wb.save(OUTPUT_FILE)

    header = (f"좌표 변환 결과 (v2)\n{'='*60}\n"
              f"성공 {done} / 실패 {failed} / 의심 {suspect} / 건너뜀(이미 있음) {skipped}\n"
              f"{'='*60}\n\n")
    body = "\n".join(report) if report else "문제 없음\n"
    with open(REPORT_FILE, "w", encoding="utf-8") as f:
        f.write(header + body)

    print("\n" + header + body)
    print(f"저장 완료 → {OUTPUT_FILE}")
    print(f"리포트   → {REPORT_FILE}")


if __name__ == "__main__":
    main()